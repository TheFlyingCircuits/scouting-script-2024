from openpyxl import Workbook
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from majora.team import Team

_centered_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def generate_spreadsheet(team_data, rankings, filename):
    """
    Notes:
    - Take some of the qualitative data (Good/OK/Bad) and make it quantitative
        - I.e. good = 3, OK = 2
        - Averages
    - Amp vs speaker points pie chart
    - Chart of match vs points, to see how they progress over the competition
        - Slope for match vs points to see how the team got better
        - Also have separate lines for teleop vs auto, and lines for amp vs speaker
    - Add the scouter's comments, also list the scouter's name
    - Generally have more high-level statistics at the top of the sheet
        - Have the complete data hidden, have to scroll down
        - Rank all the robots
        -
    - Use statbotics API and TBA API
        - Have a cache for the statbotics data?
    - How well did their teammates do in their matches?
    - How well did their opponents do in their matches?
    - One number to rule them all?


    - 3 AI generated personalities that give a review of each robot


    """
    wb = Workbook()

    main_sheet = MainSheet(team_data, rankings)
    main_sheet.add_to_workbook(wb)

    sorted_teams = sorted(list(team_data.items()), key=lambda x: int(x[0]))

    for team_number, team in sorted_teams:
        team_sheet = TeamSheet(team_number, team)
        team_sheet.add_sheet_to_workbook(wb)

    wb.save(filename)


def get_cell_str(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


class MainSheet:

    def __init__(self, team_data: dict, rankings: dict) -> None:
        self.team_data = team_data
        self.rankings = rankings

    def add_to_workbook(self, workbook: Workbook):
        ws = workbook.active
        ws.title = "Main"  # type: ignore

        self.add_rankings(ws, (1, 1))  # type: ignore

    def add_rankings(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        for column, (category, ranks) in enumerate(self.rankings.items()):
            column_title_cell_str = get_cell_str(top_left_col + column, top_left_row)
            ws[column_title_cell_str] = category
            ws[column_title_cell_str].alignment = _centered_alignment

            for row, (team_number, _) in enumerate(ranks, 1):
                cell_str = get_cell_str(top_left_col + column, top_left_row + row)
                ws[cell_str] = int(team_number)  # type: ignore


class TeamSheet:

    def __init__(self, team_number: int, team: Team) -> None:
        self.team_number = team_number
        self.team = team

    def add_sheet_to_workbook(self, workbook: Workbook):
        ws = workbook.create_sheet(str(self.team_number))
        self.add_match_data(ws, (1, 40))
        self.add_rankings(ws, (1, 1))
        self.add_ratings(ws, (4, 1))
        self.add_notes_per_match(ws, (4, 8))
        self.add_amp_vs_speaker(ws, (13, 8))

    def add_match_data(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        column_titles = [
            "Match Number",
            "Left Starting Zone",
            "Amp (Notes)",
            "Speaker (Notes)",
            "Amp (Notes)",
            "Speaker (Notes)",
            "Parked",
            "Onstage without Harmony",
            "Onstage with Harmony",
            "Notes in Trap",
            "Defense",
            "Auto",
            "Speed",
            "Pickup",
            "Scoring",
            "Driver",
            "Balance",
            "Pick",
            "Broken",
            "Details",
        ]

        auto_cols = [2, 3]
        teleop_cols = [4, 5]
        ratings_cols = [col for col in range(10, 18)]
        comments_cols = [18, 19]

        col_groups_with_headers = {
            "Auto": auto_cols,
            "Teleop": teleop_cols,
            "Ratings": ratings_cols,
            "Comments": comments_cols
        }

        # Write column headers
        for col_index, col_title in enumerate(column_titles, top_left_col):
            col_offset = col_index - top_left_col

            if any([col_offset in col_group for col_group in col_groups_with_headers.values()]):
                cell_str = get_cell_str(col_index, top_left_row + 1)
                ws[cell_str] = col_title
                ws[cell_str].alignment = _centered_alignment
            else:
                ws.merge_cells(start_row=top_left_row, start_column=col_index,
                               end_row=top_left_row+1, end_column=col_index)
                cell_str = get_cell_str(col_index, top_left_row)
                ws[cell_str] = col_title
                ws[cell_str].alignment = _centered_alignment

        for group_name, col_group in col_groups_with_headers.items():
            ws.merge_cells(start_row=top_left_row, start_column=top_left_col + col_group[0],
                           end_row=top_left_row, end_column=top_left_col + col_group[-1])
            cell_str = get_cell_str(top_left_col + col_group[0], top_left_row)
            ws[cell_str] = group_name
            ws[cell_str].alignment = _centered_alignment

        # Write actual data
        rows = []
        for match in self.team.matches:
            rows.append(match.to_list())

        for row_index, row in enumerate(rows, top_left_row + 2):
            for col_index, cell_data in enumerate(row, top_left_col):
                cell_str = get_cell_str(col_index, row_index)
                ws[cell_str] = cell_data

    def add_rankings(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        category_title_cell_str = get_cell_str(top_left_col, top_left_row)
        rank_title_cell_str = get_cell_str(top_left_col + 1, top_left_row)
        ws[category_title_cell_str] = "Category"
        ws[rank_title_cell_str] = "Rank"

        for index, (category, rank) in enumerate(self.team.stats.rankings.items(), 1):
            category_cell_str = get_cell_str(top_left_col, top_left_row + index)
            rank_cell_str = get_cell_str(top_left_col + 1, top_left_row + index)

            ws[category_cell_str] = category
            ws[rank_cell_str] = rank

    def add_ratings(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        ratings_columns = {
            "Auto": "avg_rating_auto",
            "Speed": "avg_rating_speed",
            "Pickup": "avg_rating_pickup",
            "Scoring": "avg_rating_scoring",
            "Driver": "avg_rating_driver",
            "Balance": "avg_rating_balance",
            "Pick": "avg_rating_pick",
        }

        ws.merge_cells(start_row=top_left_row, start_column=top_left_col,
                       end_row=top_left_row, end_column=top_left_col + len(ratings_columns) - 1)
        title_bar_cell_str = get_cell_str(top_left_col, top_left_row)
        ws[title_bar_cell_str] = "Average Ratings"
        ws[title_bar_cell_str].alignment = _centered_alignment

        for index, (col_title, attr) in enumerate(ratings_columns.items()):
            title_cell_str = get_cell_str(top_left_col + index, top_left_row + 1)
            ws[title_cell_str] = col_title
            ws[title_cell_str].alignment = _centered_alignment
            value_cell_str = get_cell_str(top_left_col + index, top_left_row + 2)
            ws[value_cell_str] = round(getattr(self.team, attr), 2)

    def add_notes_per_match(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        titles = [
            "Match Number",
            "Total",
            "Auto",
            "Tele",
        ]

        for index, title in enumerate(titles):
            title_cell_str = get_cell_str(top_left_col + index, top_left_row)
            ws[title_cell_str] = title
            ws[title_cell_str].alignment = _centered_alignment

        for row, match in enumerate(self.team.matches, 1):
            match_number_cell_str = get_cell_str(top_left_col, top_left_row + row)
            ws[match_number_cell_str] = row  # type: ignore

            attrs = [
                "total_notes",
                "total_notes_auto",
                "total_notes_tele",
            ]

            for col, attr in enumerate(attrs, 1):
                cell_str = get_cell_str(top_left_col + col, top_left_row + row)
                ws[cell_str] = getattr(match, attr)

        chart = LineChart()
        chart.title = "Notes Per Match"
        chart.y_axis.title = "Notes"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 30

        data = Reference(ws, min_col=top_left_col + 1, max_col=top_left_col + len(titles) - 1,
                         min_row=top_left_row, max_row=top_left_row + len(self.team.matches))
        chart.add_data(data, titles_from_data=True)

        ws.add_chart(chart, get_cell_str(top_left_col, top_left_row))

    def add_amp_vs_speaker(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        top_left_col, top_left_row = top_left_cell

        amp_title_cell_str = get_cell_str(top_left_col, top_left_row)
        ws[amp_title_cell_str] = "Amp Notes"
        amp_value_cell_str = get_cell_str(top_left_col + 1, top_left_row)
        ws[amp_value_cell_str] = self.team.total_notes_amp  # type: ignore

        speaker_title_cell_str = get_cell_str(top_left_col, top_left_row + 1)
        ws[speaker_title_cell_str] = "Speaker Notes"
        speaker_value_cell_str = get_cell_str(top_left_col + 1, top_left_row + 1)
        ws[speaker_value_cell_str] = self.team.total_notes_speaker  # type: ignore

        chart = PieChart()
        chart.title = "Amp vs. Speaker"

        data = Reference(ws, min_col=top_left_col + 1, max_col=top_left_col + 1,
                         min_row=top_left_row, max_row=top_left_row + 1)
        labels = Reference(ws, min_col=top_left_col, max_col=top_left_col,
                           min_row=top_left_row, max_row=top_left_row + 1)

        chart.add_data(data)
        chart.set_categories(labels)

        ws.add_chart(chart, get_cell_str(top_left_col, top_left_row))

    def add_pit_scouting(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        # TODO:
        ...

    def add_defense(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_council(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_statbotics(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_tba(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_teammate_quality(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_opponent_quality(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...

    def add_comments_summary(self, ws: Worksheet, top_left_cell: tuple[int, int]):
        ...
